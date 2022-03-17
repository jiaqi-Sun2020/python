import requests
import json
import os
import time
from openpyxl import load_workbook
import numpy as np

global filename
filename = "list.xlsx"
headers = {
        'apptype': '1',
        'Content-Type':'application/json',
        'authority': 'api.youpin898.com',
        'method': 'POST',
        'path': '/api/homepage/es/commodity/GetCsGoPagedList',
        'origin': 'https://www.youpin898.com' ,
        'referer': 'https://www.youpin898.com/' ,
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36 Edg/97.0.1072.62'
    }
url = 'https://api.youpin898.com/api/homepage/es/commodity/GetCsGoPagedList'


def get_price_1(Payload):  #长租价格
    #反爬机制是UA检测:(门户网站服务器会检测对应请求的载体身份标识,为某一个浏览器(正常请求,服务器允访问),为某一爬虫(不正常,服务器拒绝问))
    #UA伪装:伪装成浏览器
    headers = {
        'apptype': '1',
        'Content-Type': 'application/json',
        'authority': 'api.youpin898.com',
        'method': 'POST',
        'path': '/api/homepage/es/commodity/GetCsGoPagedList',
        'origin': 'https://www.youpin898.com',
        'referer': 'https://www.youpin898.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36 Edg/97.0.1072.62'
    }
    url = 'https://api.youpin898.com/api/homepage/es/commodity/GetCsGoPagedList'
    #===================================
    #post

    Payload = json.dumps(Payload)
    #print(Payload)
    response = requests.post(url=url, data = Payload, headers=headers)
    #改成json文件
    obj = response.json()
    #print(obj[str('Data')][str('CommodityList')])#获取到了json里面Msg的文件

    #持久化存储(存储了获取的json文件)
    with open('./道具信息_1.json','w',encoding= 'utf-8') as fp:
        json.dump(obj,fp = fp, ensure_ascii = False)

    #================================================================下面对获取的json文件进行分析以及数据的整理
    #所有价格信息目录
    XXX = obj[str('Data')][str('CommodityList')]
    LongLeaseUnitPrice = []
    name = []
    if XXX !=None:    #如果存在数据的话
        for T in  XXX:
            LongLeaseUnitPrice.append(float(T[str('LongLeaseUnitPrice')]) )#字符串要转成浮点型才能放入min内比较
            name.append(T[str('UserNickName')])

        #print(LongLeaseUnitPrice)

        min_long = min(list(LongLeaseUnitPrice))
        min_name  = np.argmin(list(LongLeaseUnitPrice))
        name_id = name[min_name]
        leasename = obj[str('Data')][str('CommodityList')][0][str('CommodityName')]
        #数据大小
        #length = len(obj[str('Data')][str('CommodityList')])
        #print("数据大小{}".format(length))
        #print(leasename)
        #print('今日最小长租价格为" {price} ",是" {id} "出租的'.format(price = min_long, id = name_id))#显示出最低长租价格并且表明出租人是谁
        return min_long , name_id , leasename
    else:
        #print('TemplateInfo' in  obj[('Data')])
        #if obj[('Data')][str('TemplateInfo')][str("CommodityName")] != None:  # 如果存在数据的话
        if 'TemplateInfo' in  obj[('Data')]:

            #print(obj[('Data')][str('TemplateInfo')])
            if obj[('Data')][str('TemplateInfo')] != None:
                leasename = obj[('Data')][str('TemplateInfo')][str("CommodityName")]
            else:
                leasename = None
        else:
            leasename = None
        min_long=name_id=None

        return min_long , name_id , leasename

def get_price_2(Payload):  #最贵价格
    headers = {
        'apptype': '1',
        'Content-Type': 'application/json',
        'authority': 'api.youpin898.com',
        'method': 'POST',
        'path': '/api/youpin/commodity/purchase/find',
        'origin': 'https://www.youpin898.com',
        'referer': 'https://www.youpin898.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36 Edg/97.0.1072.62'
    }
    url = 'https://api.youpin898.com/api/youpin/commodity/purchase/find'
    Payload = json.dumps(Payload)
    # print(Payload)
    response = requests.post(url=url, data=Payload, headers=headers)
    # 改成json文件
    obj = response.json()
    # print(obj[str('Data')][str('CommodityList')])#获取到了json里面Msg的文件

    # print(obj)
    # 持久化存储(存储了获取的json文件)
    with open('./道具信息_2.json', 'w', encoding='utf-8') as fp:
        json.dump(obj, fp=fp, ensure_ascii=False)

    buy_price = []
    buy_name = []
    XXX = obj[str('data')][str('response')]

    # 根据获取的数据来得到出租的人和物
    if len(XXX) !=0:
        for T in XXX:
            buy_name.append(T[str('userNickname')])  # 获取名字
            buy_price.append(float(T[str('unitPrice')] / 100))
        max_buyprice = max(buy_price)
        buyname = buy_name[np.argmax(buy_price)]
        return max_buyprice, buyname
    else:
        buyname = max_buyprice = None
        return max_buyprice, buyname



#接下来对获取时间戳
def get_time():
    t = time.localtime()
    timecellect = [t.tm_year, t.tm_mon, t.tm_mday,t.tm_hour, t.tm_min, t.tm_sec]
    nowtime = "{}/{}/{}".format(t.tm_hour, t.tm_min, t.tm_sec)
    nyr = "{}-{}-{}".format(t.tm_year, t.tm_mon, t.tm_mday)
    return nowtime,nyr

#xlwt编写excel
def excel_init():
    nowtime, nyr = get_time()
    if os.path.exists(filename) == False:
        print("no exists file")
    else:
        print("exists file")

def write_excel_xls_append(path,price,name,max_price,row):
    wb = load_workbook(path,data_only = False)
    #ws = wb.get_sheet_by_name("租金")
    ws = wb["list"]
    n = ws.cell(row = row , column = 4)#名字列
    p = ws.cell(row=row, column = 5)#租金列
    q = ws.cell(row=row, column=6)  # 价格列
    n.value = name
    p.value = price# 价格列
    q.value = max_price
    wb.save(path)

#主函数
def main():
    #post请求包
    pagesize = 100
    Payload_1 =[]


    min_len =43951                                           #==================================================================================在这里改在id起始
    max_len =44000                                       #==================================================================================在这里改在id结束
    for z in range(min_len,max_len+1):
       c =  {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "{}".format(z)}
       Payload_1.append(c)

    excel_init()#初始化表格
    while(1):
        nowtime,nyr= get_time()
        print(nowtime)
        p =0
        for i in range(min_len ,max_len+1):
            Paylod_1 = {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1,'templateId': "{}".format(i)}
            Paylod_2 = Payload ={'pageIndex': 1, 'pageSize': pagesize,'templateId': "{}".format(i)}
            min_long, name_id, leasename = get_price_1(Paylod_1)  # 获取租金
            buyname = max_buyprice=None
            max_buyprice, buyname = get_price_2(Paylod_2)      #获取价格
            #写入excel
            #编辑
            if min_long !=None  and  buyname !=None:
                if min_long>1:
                    write_excel_xls_append(filename,min_long,leasename,max_buyprice,p+1+1)
                    p=p+1
            print('现在{leasename}最小长租价格为" {min_long} ",是" {name_id} "出租的 // {buyname}最高求购价格为 {max_buyprice}'.format(min_long=min_long, name_id=name_id, leasename=leasename, buyname=buyname,max_buyprice=max_buyprice))  # 显示出最低长租价格并且表明出租人是谁
        print(20*"*"+"over")
        time.sleep(100000)#=====================================在此改间隔爬取时间单位为秒


if __name__ =="__main__":
    main()