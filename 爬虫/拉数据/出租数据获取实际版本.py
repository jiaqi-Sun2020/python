import requests
import json
import os
import time
from openpyxl import load_workbook
import numpy as np
import gc

global filename
filename = "./出租.xlsx"
headers = {
        'apptype': '1',
        'Content-Type':'application/json',
        'authority': 'api.youpin898.com',
        'method': 'POST',
        'path': '/api/homepage/es/commodity/GetCsGoPagedList',
        'origin': 'https://www.youpin898.com' ,
        'referer': 'https://www.youpin898.com/' ,
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36 Edg/97.0.1072.62'
    }
url = 'https://api.youpin898.com/api/homepage/es/commodity/GetCsGoPagedList'
Payload = {'templateId': "892", 'pageSize': 30, 'pageIndex': 1, 'sortType': 2, 'listSortType': 3, 'listType': 30}
def get_price(headers,url ,Payload):
    #反爬机制是UA检测:(门户网站服务器会检测对应请求的载体身份标识,为某一个浏览器(正常请求,服务器允访问),为某一爬虫(不正常,服务器拒绝问))
    #UA伪装:伪装成浏览器

    #===================================
    #post

    Payload = json.dumps(Payload)
    #print(Payload)
    response = requests.post(url=url, data = Payload, headers=headers)
    #改成json文件
    obj = response.json()
    #print(obj[str('Data')][str('CommodityList')])#获取到了json里面Msg的文件

    #持久化存储(存储了获取的json文件)
    with open('./道具信息.json','w',encoding= 'utf-8') as fp:
        json.dump(obj,fp = fp, ensure_ascii = False)

    #================================================================下面对获取的json文件进行分析以及数据的整理
    #所有价格信息目录
    XXX = obj[str('Data')][str('CommodityList')]
    LongLeaseUnitPrice = []
    name = []
    for T in  XXX:
        LongLeaseUnitPrice.append(float(T[str('LongLeaseUnitPrice')]) )#字符串要转成浮点型才能放入min内比较
        name.append(T[str('UserNickName')])

    #print(LongLeaseUnitPrice)

    min_long = min(list(LongLeaseUnitPrice))
    min_name  = np.argmin(list(LongLeaseUnitPrice))
    name_id = name[min_name]
    shellname = obj[str('Data')][str('CommodityList')][0][str('CommodityName')]
    #数据大小
    length = len(obj[str('Data')][str('CommodityList')])
    #print("数据大小{}".format(length))
    #print(shellname)
    #print('今日最小长租价格为" {price} ",是" {id} "出租的'.format(price = min_long, id = name_id))#显示出最低长租价格并且表明出租人是谁
    return min_long , name_id , shellname




#接下来对获取时间戳
def get_time():
    t = time.localtime()
    timecellect = [t.tm_year, t.tm_mon, t.tm_mday,t.tm_hour, t.tm_min, t.tm_sec]
    nowtime = "{}/{}/{}".format(t.tm_hour, t.tm_min, t.tm_sec)
    nyr = "{}-{}-{}".format(t.tm_year, t.tm_mon, t.tm_mday)
    return nowtime,nyr

#xlwt编写excel
def excel_init():
    nowtime, nyr = get_time()

    if os.path.exists(filename) == False:
        print(" no exists file")
    else:
        print("exists file")

def write_excel_xls_append(ws,price,name,row):
    n = ws.cell(row = row , column = 3)#名字列
    p = ws.cell(row=row, column = 5)#价格列
    n.value = name
    p.value =  price# 价格列


#主函数
def main():
    headers = {
        'apptype': '1',
        'Content-Type': 'application/json',
        'authority': 'api.youpin898.com',
        'method': 'POST',
        'path': '/api/homepage/es/commodity/GetCsGoPagedList',
        'origin': 'https://www.youpin898.com',
        'referer': 'https://www.youpin898.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36 Edg/97.0.1072.62'
    }
    url = 'https://api.youpin898.com/api/homepage/es/commodity/GetCsGoPagedList'

    #post请求包
    pagesize = 100
    Payload =[
        {'listSortType': 3,'listType': 30,'pageIndex': 1,'pageSize': pagesize,'sortType': 1,'templateId': "43951"},#1原版齿锯
        {'listSortType': 3,'listType': 30,'pageIndex': 1,'pageSize': pagesize,'sortType': 1,'templateId': "49059"}, #2渐变大理石
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "2791"},#3原版短剑
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "43399"},#4火沙鹰
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "741"},#5酒精双栖
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "1678"},#6原版熊刀
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "44488"},#7原版刺刀
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "44829"},#8原版爪子刀
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "45197"},#9原版M9
        {'templateId': "892", 'pageSize': pagesize, 'pageIndex': 1, 'sortType': 2, 'listSortType': 3, 'listType': 30},# 10蝴蝶刀
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "743"},# 11略磨火神ak
        {'templateId': "2408", 'pageSize': pagesize, 'pageIndex': 1, 'sortType': 2, 'listSortType': 3, 'listType': 30}, # 12崭新二号玩家
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "43766"},# 13专业手套（★） | 大腕 (久经沙场)
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "46003"},#14专业手套（★） | 一线特工 (久经沙场)
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "1944"},#15AWP | 野火 (略有磨损)
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "51135"},#16运动手套（★） | 夜行衣 (久经沙场)
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "34833"},#17 M4A1 消音型 | 印花集 (久经沙场)
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "1644"},#18 沙漠之鹰 | 印花集 (久经沙场)
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "407"},#19 AWP | 二西莫夫 (久经沙场)
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "12359"},#20 蝴蝶刀（★） | 北方森林 (久经沙场)
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "2130"},#21 专业手套（★） | 渐变之色 (久经沙场)
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "541"},#22 AWP | 野火 (久经沙场)
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "43393"},#23 AWP | 雷击 (崭新出厂)
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "46877"},#24 海豹短刀（★）
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "4298"},#25 弯刀（★）
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "44787"},#26 系绳匕首（★）
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "2097"},#27 裹手（★） | 钴蓝骷髅 (久经沙场)
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "61735"},#28 格洛克 18 型 | 伽玛多普勒 (崭新出厂)
        {'listSortType': 3, 'listType': 30, 'pageIndex': 1, 'pageSize': pagesize, 'sortType': 1, 'templateId': "1986"},#29 M9 刺刀（★） | 澄澈之水 (久经沙场)










    ]
    nowtime, nyr = get_time()

    excel_init()#初始化表格
    #打开excel表格
    start = time.time()
    wb = load_workbook(filename, data_only=False)
    # ws = wb.get_sheet_by_name("租金")
    ws = wb["租金"]
    end_excel = time.time()
    print('打开完毕,花费时间{}'.format(end_excel - start))

    while(1):
        nowtime,nyr= get_time()
        print(nowtime)
        for i in range(len(Payload)):
            min_long, name_id, shellname = get_price(headers,url,Payload[i] )
            print('现在{shellname}最小长租价格为" {price} ",是" {id} "出租的'.format(price = min_long, id = name_id, shellname = shellname))#显示出最低长租价格并且表明出租人是谁
            #写入excel
            #编辑
            write_excel_xls_append(ws,min_long,shellname,i+1+2)


        #关闭表格
        wb.save(filename)
        wb.close()  # 对程序中只读的workbook
        del wb, ws
        gc.collect()
        # 清除内存等
        end_all = time.time()
        print('总花费时间{}'.format(end_all - start))
        time.sleep(5)#=====================================在此改间隔爬取时间单位为秒


if __name__ =="__main__":
    main()