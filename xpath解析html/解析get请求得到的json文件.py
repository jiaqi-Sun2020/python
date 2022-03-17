import requests
import json
import gc
from openpyxl import load_workbook
tag_ids = None
import time

#获取json数据的函数
def get_json_data(id,tag_ids):

    headers = {
        'referer': 'https://buff.163.com/goods/871196',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36 Edg/98.0.1108.62',
        'cookie': '_ntes_nnid=273aa6cd2c9b4834de70916de9b28928,1636197670821; UM_distinctid=17cf4fb278885-029e8a313b0acb-561a135a-1fa400-17cf4fb2789f88; _ntes_nuid=273aa6cd2c9b4834de70916de9b28928; unisdk_udid=71d4d9f8b13b373044a18cf3a46299b0; NTES_P_UTID=HQOJqyJodQV7kZEg99eDIWCxetfxGnhd|1639838795; nts_mail_user=sjq1623958071@163.com:-1:1; timing_user_id=time_ysX566F1ZC; Device-Id=S93qHFLLAFCsOwFsn3gZ; _ga=GA1.2.1996916940.1646207084; P_INFO=15306260007|1646399177|1|netease_buff|00&99|null&null&null#jis&320500#10#0|&0|null|15306260007; remember_me=U1101189625|31rgSZ68DsNQzjUN6DbEI10x7WkGt3k7; Locale-Supported=zh-Hans; game=csgo; _gid=GA1.2.466422361.1646583975; session=1-s4AgkVO-iVRAV82ZnFCkL2Talpr4gWSoo6oUqBPf2ll_2037141153; _gat_gtag_UA_109989484_1=1; csrf_token=IjFhNzk5MTIyNDMwZjMzMGEzOWE5NjUxZGRiNzJlYjNjMjhjYTM3ZmMi.FQZySw.8GGiJ3fjqPrMeGiX9QaMjTjkY2I',

    }
    url = 'https://buff.163.com/api/market/goods/sell_order'

    param = {
        'game': 'csgo',
        'goods_id': str(id),
        'page_num': '1',
        'sort_by': 'default',
        'mode': '',
        'allow_tradable_cooldown': '1',
        '_': '1646307155097',
        #大部分网页要登录的话是得用cookie的这就是你的用户名!!!!!

    }
    if tag_ids !=None:
        param['tag_ids'] = str(tag_ids)
    #print(param)
    text = []
    while(len(text) <= 3000):
        get_data = requests.get(url=url,params=param,headers = headers)#获取请求文件

        text = get_data.text#转化为text查看是否存在空的字符
        #print(text)
        if len(text) >= 3000:
            print("yes")
            break
        else:
            print('No')
            #print(get_data.json())

        time.sleep(1)
    #print(get_data)
    #使用utf-8解码
    get_data.content.decode("utf-8")


    obj = get_data.json()

    #print(obj)

    with open('./数据暂存.json','w',encoding= 'utf-8') as fp:
        json.dump(obj,fp = fp, ensure_ascii = False)
    return obj

#获取价格函数
def get_price(id,tag_ids):
    #对data解包
    json_data = get_json_data(id=id,tag_ids=tag_ids)#获取json数据文件


    allsheller_data = json_data['data']['items']    #所有出售人的数据
    #print(json_data['data']['items'])
    #if 'items' in json_data['data'].keys():
    if len(allsheller_data)>0 and json_data !=None:
        price = float(allsheller_data[0]['price'])
        goods_infos=json_data['data']['goods_infos']
        good_name  = goods_infos[str(id)]['name']
        #print(good_name)
        return id,good_name,price
    else:
        return id,None,None

def write_excel_xls_append(ws,price,name,row):
    #n = ws.cell(row = row , column = 3)#名字列
    p = ws.cell(row=row, column = 8)#价格列
    #n.value = name
    p.value =  price# 价格列



path = './2M.xlsx'
#主函数====
while(1):
    row = 0
    #准备打开excel
    start = time.time()
    print('打开excel')
    wb = load_workbook(path, data_only=False)
    # ws = wb.get_sheet_by_name("租金")
    ws = wb["租金"]
    end_excel = time.time()
    print('打开完毕,花费时间{}'.format(end_excel-start))
    #打开完毕
    #使用时物品的id是前面的一个，涂装id是后面一个
    # goods_id = [
    #     (759326,None),#原版锯齿（★）
    #     (871196,10199483),#格洛克 18 型 | 伽玛多普勒 (崭新出厂)
    #     (43074,None),#M9 刺刀（★） | 澄澈之水 (久经沙场)
    # ]  #这些都是商品的id
    goods_id = [
        (759326, None),  # 原版锯齿（★）
        (835939, None),  # 专业手套（★） | 渐变大理石 (久经沙场)
        (759509, None),  # 短剑（★）
        (34389, None),  # 沙漠之鹰 | 炽烈之炎 (崭新出厂)
        (45451, None),  # 运动手套（★） | 双栖 (久经沙场)
        (759460, None),  # 熊刀（★）
        (42349, None),  # 刺刀（★）
        (42961, None),  # 爪子刀（★）
        (43052, None),  # M9 刺刀（★）
        (42530, None),  # 蝴蝶刀（★
        (33976, None),  # AK-47 | 火神 (略有磨损)
        (779333, None),  # M4A1 消音型 | 二号玩家 (崭新出厂)
        (45376, None),  # 专业手套（★） | 大腕 (久经沙场)
        (836015, None),  # 专业手套（★） | 一线特工 (久经沙场)
        (773720, None),  # AWP | 野火 (略有磨损)
        (835873, None),  # 运动手套（★） | 夜行衣 (久经沙场)
        (835624, None),  # M4A1 消音型 | 印花集 (久经沙场)
        (781598, None),  # 沙漠之鹰 | 印花集 (久经沙场)
        (34066, None),  # AWP | 二西莫夫 (久经沙场)
        (42538, None),  # 蝴蝶刀（★） | 北方森林 (久经沙场)
        (45508, None),  # 专业手套（★） | 渐变之色 (久经沙场)
        (773698, None),  # AWP | 野火 (久经沙场)
        (34099, None),  # AWP | 雷击 (崭新出厂)
        (773696, None),  # 海豹短刀（★）
        (44002, None),  # 弯刀（★）
        (776671, None),  # 系绳匕首（★）
        (45471, None),  # 裹手（★） | 钴蓝骷髅 (久经沙场)
        (871196, 10199483),  # 格洛克 18 型 | 伽玛多普勒 (崭新出厂)  特殊涂装
        (43074, None),  # M9 刺刀（★） | 澄澈之水 (久经沙场)
    ]  # 这些都是商品的id
    for id,tag_ids in goods_id:
        id,good_name,price=get_price(id,tag_ids)
        str_print = r"物品名称{good_name},价格{price},物品编号{id},特殊涂装编号{tag_ids}".format(good_name=good_name,price=price,id=id,tag_ids=tag_ids)
        print(str_print)
        if good_name !=None:
            write_excel_xls_append(ws,price,good_name,row+3)
            row = row+1
        time.sleep(0)
    #保存excel
    wb.save(path)
    wb.close()  # 对程序中只读的workbook
    del wb, ws
    gc.collect()
    #清除内存等
    #time.sleep(100)
    break  #跳出while循环

end_all = time.time()

print('总花费时间{}'.format(end_all-start))
print('=======over=======')

